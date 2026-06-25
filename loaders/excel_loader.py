from openpyxl import load_workbook

def load_excel(path):

    wb = load_workbook(path)

    text = ""

    for sheet in wb.worksheets:

        for row in sheet.iter_rows(values_only=True):

            text += " ".join(
                str(cell)
                for cell in row
                if cell is not None
            )

            text += "\n"

    return text
